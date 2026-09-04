"""
05_build_excel_template.py

Builds the Excel template a person would actually use to log their
intentions and events by hand -- this is the "data collection layer" of the
project (see README section on where Excel fits in the pipeline). It is a
genuine input tool, not a report: dropdown validation keeps the categorical
fields consistent at the point of entry, which is exactly the kind of mess
this project's cleaning script (03_clean_and_load.py) has to fix when data
is instead free-typed.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

FONT_NAME = "Calibri"
HEADER_FILL = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # pale yellow = fill this in
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------------------------------------------------------------------
# SHEET 1: Instructions
# ---------------------------------------------------------------------------
ws_intro = wb.active
ws_intro.title = "Instructions"
ws_intro.sheet_view.showGridLines = False
ws_intro.column_dimensions["A"].width = 100

lines = [
    ("The Graveyard of Good Intentions — Tracking Log", 16, True),
    ("", 11, False),
    ("How to use this workbook", 13, True),
    ("1. Log every intention you create on the 'Intentions' sheet, one row each.", 11, False),
    ("2. Every time something happens to that intention (you research it, plan it, start it,", 11, False),
    ("   work on it, postpone it, abandon it, or complete it), add one row to the 'Events' sheet.", 11, False),
    ("3. Yellow cells are the ones you fill in. Use the dropdowns where provided --", 11, False),
    ("   this keeps categories consistent so the data can actually be analyzed later.", 11, False),
    ("4. An 'intention_id' links the two sheets together: give each new intention the next", 11, False),
    ("   free number, and use that same number on every Events row for that intention.", 11, False),
    ("", 11, False),
    ("Why two sheets instead of one?", 13, True),
    ("An intention is a single thing (e.g. 'Learn SQL properly'), but things happen to it", 11, False),
    ("multiple times over its life (researched twice, started, postponed, finished). Keeping", 11, False),
    ("one row per intention and a separate row per event is what lets you later calculate", 11, False),
    ("things like 'how many times did I postpone this' or 'how long did this survive'.", 11, False),
    ("", 11, False),
    ("This workbook is the raw input layer only", 13, True),
    ("It intentionally is NOT where the analysis happens. Cleaning, analysis, and", 11, False),
    ("visualization happen downstream in Python / SQL / Power BI / Tableau once you", 11, False),
    ("export these sheets to CSV. See the project README for the full pipeline.", 11, False),
]
row = 1
for text, size, bold in lines:
    cell = ws_intro.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color="1F3A5F" if bold else "000000")
    row += 1

# ---------------------------------------------------------------------------
# SHEET 2: Lookup lists (hidden-ish support sheet for dropdowns)
# ---------------------------------------------------------------------------
ws_lookup = wb.create_sheet("Lookups")
CATEGORIES = ["Career", "Learning & Skills", "Fitness & Health", "Creative", "Financial",
              "Home & Admin", "Relationships", "Travel", "Side Project", "Hobby"]
EVENT_TYPES = ["CREATED", "RESEARCHED", "PLANNED", "STARTED", "CONTINUED",
               "POSTPONED", "ABANDONED", "COMPLETED", "REPLACED"]
YES_NO = ["Yes", "No"]
SCALE_1_5 = [1, 2, 3, 4, 5]
BARRIERS = ["Lack of time", "Lack of money", "Lack of motivation", "Unclear next step",
            "Perfectionism / fear of failure", "Competing priorities", "Lost interest",
            "Blocked by dependency", "Low energy / burnout", "No accountability",
            "Overestimated effort required"]

ws_lookup["A1"] = "Category"
ws_lookup["B1"] = "EventType"
ws_lookup["C1"] = "YesNo"
ws_lookup["D1"] = "Scale1to5"
ws_lookup["E1"] = "Barrier"
for i, v in enumerate(CATEGORIES, start=2):
    ws_lookup.cell(row=i, column=1, value=v)
for i, v in enumerate(EVENT_TYPES, start=2):
    ws_lookup.cell(row=i, column=2, value=v)
for i, v in enumerate(YES_NO, start=2):
    ws_lookup.cell(row=i, column=3, value=v)
for i, v in enumerate(SCALE_1_5, start=2):
    ws_lookup.cell(row=i, column=4, value=v)
for i, v in enumerate(BARRIERS, start=2):
    ws_lookup.cell(row=i, column=5, value=v)
ws_lookup.sheet_state = "hidden"

# ---------------------------------------------------------------------------
# SHEET 3: Intentions (master log)
# ---------------------------------------------------------------------------
ws_int = wb.create_sheet("Intentions")
int_headers = [
    "intention_id", "title", "category", "motivation", "date_created",
    "perceived_importance (1-5)", "urgency (1-5)", "personal_interest (1-5)",
    "career_relevance (1-5)", "estimated_effort_hours", "has_deadline",
    "deadline_date",
]
for col, h in enumerate(int_headers, start=1):
    c = ws_int.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws_int.column_dimensions[get_column_letter(col)].width = 20
ws_int.row_dimensions[1].height = 32

example_int = [1, "Learn SQL properly", "Learning & Skills", "Career advancement",
               "2025-01-15", 4, 3, 4, 5, 20, "Yes", "2025-04-01"]
for col, v in enumerate(example_int, start=1):
    c = ws_int.cell(row=2, column=col, value=v)
    c.fill = INPUT_FILL
    c.border = BORDER
    c.font = Font(name=FONT_NAME, italic=True, color="666666")

# Data validation dropdowns for Intentions sheet (rows 2-500)
dv_category = DataValidation(type="list", formula1="=Lookups!$A$2:$A$11", allow_blank=True)
dv_yesno = DataValidation(type="list", formula1="=Lookups!$C$2:$C$3", allow_blank=True)
dv_scale = DataValidation(type="list", formula1="=Lookups!$D$2:$D$6", allow_blank=True)
for dv in (dv_category, dv_yesno, dv_scale):
    ws_int.add_data_validation(dv)
dv_category.add("C3:C500")
dv_scale.add("F3:I500")
dv_yesno.add("K3:K500")

for row_i in range(3, 501):
    for col_i in range(1, len(int_headers) + 1):
        ws_int.cell(row=row_i, column=col_i).border = BORDER
ws_int.freeze_panes = "A2"

# ---------------------------------------------------------------------------
# SHEET 4: Events (log of what happened, one row per event)
# ---------------------------------------------------------------------------
ws_evt = wb.create_sheet("Events")
evt_headers = ["log_id", "intention_id", "event_date", "event_type",
               "hours_spent", "barrier (if postponed/abandoned)", "note"]
for col, h in enumerate(evt_headers, start=1):
    c = ws_evt.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws_evt.column_dimensions[get_column_letter(col)].width = 22
ws_evt.row_dimensions[1].height = 32

example_events = [
    [1, 1, "2025-01-15", "CREATED", 0, "", ""],
    [2, 1, "2025-01-18", "RESEARCHED", 1.5, "", "Looked at a couple of course options"],
    [3, 1, "2025-01-25", "STARTED", 0, "", "Signed up for a course"],
    [4, 1, "2025-02-10", "POSTPONED", 0, "Lack of time", "Busy at work this week"],
]
for r_offset, row_vals in enumerate(example_events):
    r = r_offset + 2
    for col, v in enumerate(row_vals, start=1):
        c = ws_evt.cell(row=r, column=col, value=v)
        c.fill = INPUT_FILL
        c.border = BORDER
        c.font = Font(name=FONT_NAME, italic=True, color="666666")

dv_event_type = DataValidation(type="list", formula1="=Lookups!$B$2:$B$10", allow_blank=True)
dv_barrier = DataValidation(type="list", formula1="=Lookups!$E$2:$E$12", allow_blank=True)
ws_evt.add_data_validation(dv_event_type)
ws_evt.add_data_validation(dv_barrier)
dv_event_type.add("D6:D1000")
dv_barrier.add("F6:F1000")

for row_i in range(6, 1001):
    for col_i in range(1, len(evt_headers) + 1):
        ws_evt.cell(row=row_i, column=col_i).border = BORDER
ws_evt.freeze_panes = "A2"

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
wb.save("excel/intention_tracker_template.xlsx")
print("Saved excel/intention_tracker_template.xlsx")
